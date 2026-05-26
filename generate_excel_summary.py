"""
generate_excel_summary.py — DDCSE v3.0 Stakeholder Excel Export
================================================================
Produces a clean, formatted .xlsx workbook for recruiters and
non-technical stakeholders. No Python required to open.

Sheets:
  1. Portfolio Dashboard  — one-row-per-company summary with RAG status
  2. Covenant Detail      — all covenant tests across all 8 companies
  3. Debt Stack           — full tranche-level debt breakdown
  4. Stress Scenarios     — key macro stress NLR outcomes per company
  5. About                — engine description, sources, disclaimer
"""

import sys, os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_NUMBER_00

sys.path.insert(0, os.path.dirname(__file__))
from real_cases import CASE_REGISTRY, list_cases
from analytics_v2 import evaluate_package, portfolio_severity_score, macro_shock_matrix

# ─── Palette ──────────────────────────────────────────────────────────────────
NAVY      = "0A1628"
BLUE      = "1A3A5C"
MIDBLUE   = "2C5282"
LGRAY     = "F4F5F7"
MGRAY     = "8A9BB0"
DGRAY     = "2C3E50"
RED       = "C0392B"
RED_LIGHT = "F9E5E4"
AMBER     = "D4680A"
AMB_LIGHT = "FDF3E7"
GREEN     = "1A7A4A"
GRN_LIGHT = "E8F5EE"
WHITE     = "FFFFFF"
CREAM     = "FAFBFD"
NEWBLUE   = "EEF4FF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DGRAY, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def border_thin():
    s = Side(style="thin", color="D0D7E2")
    return Border(left=s, right=s, top=s, bottom=s)

def border_medium():
    s = Side(style="medium", color="8A9BB0")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def status_fill(status):
    return {
        "BREACH":    fill(RED_LIGHT),
        "WATCHLIST": fill(AMB_LIGHT),
        "WARNING":   fill(AMB_LIGHT),
        "COMPLIANT": fill(GRN_LIGHT),
    }.get(status, fill(LGRAY))

def status_font(status, bold=True):
    color = {
        "BREACH":    RED,
        "WATCHLIST": AMBER,
        "WARNING":   AMBER,
        "COMPLIANT": GREEN,
    }.get(status, DGRAY)
    return font(bold=bold, color=color)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def header_row(ws, row_num, headers, col_widths=None):
    """Write a navy header row."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=c, value=h)
        cell.fill = fill(NAVY)
        cell.font = font(bold=True, color=WHITE, size=9)
        cell.alignment = center()
        cell.border = border_thin()
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w

def title_block(ws, title, subtitle, row=1):
    """Write a title + subtitle block at top of sheet."""
    ws.row_dimensions[row].height = 28
    ws.row_dimensions[row+1].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    t = ws.cell(row=row, column=1, value=title)
    t.fill = fill(NAVY)
    t.font = font(bold=True, color=WHITE, size=14)
    t.alignment = left_mid()
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+1, end_column=12)
    s = ws.cell(row=row+1, column=1, value=subtitle)
    s.fill = fill(BLUE)
    s.font = font(color="B0C4DE", size=9)
    s.alignment = left_mid()
    # amber accent
    ws.row_dimensions[row+2].height = 4
    for c in range(1, 13):
        ws.cell(row=row+2, column=c).fill = fill(AMBER)
    return row + 3  # next usable row


# ─── SHEET 1: Portfolio Dashboard ─────────────────────────────────────────────
def build_dashboard(wb):
    ws = wb.create_sheet("📊 Portfolio Dashboard")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"

    next_row = title_block(
        ws,
        "DDCSE v3.0 — Dynamic Debt Covenant Surveillance Engine",
        f"Portfolio Dashboard  ·  8 Credits  ·  6 Sectors  ·  Generated {date.today().strftime('%B %d, %Y')}"
    )

    # KPI summary row
    cases = list_cases()
    breaches  = sum(1 for c in cases if c["nlr"] > c["nlr_threshold"])
    watchlist = sum(1 for c in cases if c["nlr"] <= c["nlr_threshold"]
                    and (c["nlr_threshold"] - c["nlr"]) / c["nlr_threshold"] < 0.10)
    compliant = len(cases) - breaches - watchlist
    total_debt_bn = sum(CASE_REGISTRY[c["ticker"]]["total_debt"] for c in cases) / 1000

    kpis = [
        ("Credits Monitored", len(cases), NAVY),
        ("Covenant Breaches", breaches, RED),
        ("Watchlist", watchlist, AMBER),
        ("Compliant", compliant, GREEN),
        ("Total Debt (USD)", f"${total_debt_bn:.1f}B", MIDBLUE),
    ]
    ws.row_dimensions[next_row].height = 30
    ws.row_dimensions[next_row+1].height = 18
    for i, (label, val, color) in enumerate(kpis):
        col = i * 2 + 1
        ws.merge_cells(start_row=next_row, start_column=col, end_row=next_row, end_column=col+1)
        v = ws.cell(row=next_row, column=col, value=val)
        v.fill = fill(color)
        v.font = font(bold=True, color=WHITE, size=16)
        v.alignment = center()
        ws.merge_cells(start_row=next_row+1, start_column=col, end_row=next_row+1, end_column=col+1)
        l = ws.cell(row=next_row+1, column=col, value=label)
        l.fill = fill(color)
        l.font = font(color="B0C4DE", size=8)
        l.alignment = center()

    next_row += 3

    # Main table
    headers = [
        "Ticker", "Company", "Sector", "Rating", "Outlook",
        "NLR Actual", "NLR Threshold", "Buffer %", "ICR",
        "FCCR", "Severity /100", "Overall Status", "New Sector?"
    ]
    col_widths = [8, 28, 32, 8, 10, 12, 13, 10, 8, 8, 13, 16, 12]
    header_row(ws, next_row, headers, col_widths)
    next_row += 1

    new_sectors = {"BHC", "M", "LUMN"}

    for c in cases:
        case = CASE_REGISTRY[c["ticker"]]
        results = evaluate_package(case["covenants"])
        sev = portfolio_severity_score(results)
        fccr_r = next((r for r in results if r.cov_type == "fccr"), None)
        fccr_v = fccr_r.actual if fccr_r else None

        nlr = c["nlr"]
        thresh = c["nlr_threshold"]
        buf_pct = (thresh - nlr) / thresh * 100

        overall = ("BREACH" if any(r.status == "BREACH" for r in results) else
                   "WATCHLIST" if any(r.status in ("WATCHLIST", "WARNING") for r in results)
                   else "COMPLIANT")

        is_new = c["ticker"] in new_sectors
        row_fill = fill(NEWBLUE) if is_new else fill(WHITE)
        alt_fill = fill("E6EFFF") if is_new else fill(LGRAY)

        row_data = [
            c["ticker"], case["name"], case["sector"].split("—")[0].strip(),
            case["credit_rating"], case["outlook"],
            round(nlr, 2), thresh, round(buf_pct, 1), round(c["icr"], 2),
            round(fccr_v, 2) if fccr_v else "N/A",
            round(sev, 0), overall,
            "★ NEW" if is_new else "—"
        ]

        ws.row_dimensions[next_row].height = 20
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col_idx, value=val)
            cell.border = border_thin()
            cell.alignment = center() if col_idx != 3 else left_mid()

            # Status colouring
            if col_idx == 12:
                cell.fill = status_fill(overall)
                cell.font = status_font(overall)
            elif col_idx == 8:
                if buf_pct < 0:
                    cell.fill = fill(RED_LIGHT); cell.font = font(bold=True, color=RED)
                elif buf_pct < 10:
                    cell.fill = fill(AMB_LIGHT); cell.font = font(bold=True, color=AMBER)
                else:
                    cell.fill = fill(GRN_LIGHT); cell.font = font(bold=True, color=GREEN)
            elif col_idx == 11:
                sev_color = RED if sev >= 70 else (AMBER if sev >= 40 else GREEN)
                cell.fill = fill(RED_LIGHT if sev >= 70 else (AMB_LIGHT if sev >= 40 else GRN_LIGHT))
                cell.font = font(bold=True, color=sev_color)
            elif col_idx == 13 and is_new:
                cell.fill = fill("C8DBFF")
                cell.font = font(bold=True, color=BLUE)
            else:
                cell.fill = row_fill if next_row % 2 == 0 else alt_fill
                cell.font = font(bold=(col_idx == 1))

        next_row += 1

    # Legend
    next_row += 1
    legend = [("BREACH", RED_LIGHT, RED), ("WATCHLIST", AMB_LIGHT, AMBER),
              ("COMPLIANT", GRN_LIGHT, GREEN), ("★ NEW Sector", NEWBLUE, BLUE)]
    for i, (label, bg, fg) in enumerate(legend):
        col = i * 3 + 1
        c1 = ws.cell(row=next_row, column=col, value=f"■  {label}")
        c1.fill = fill(bg)
        c1.font = font(bold=True, color=fg, size=9)
        c1.alignment = center()
        c1.border = border_thin()

    ws.print_area = f"A1:{get_column_letter(len(headers))}{next_row}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1


# ─── SHEET 2: Covenant Detail ──────────────────────────────────────────────────
def build_covenant_detail(wb):
    ws = wb.create_sheet("🔍 Covenant Detail")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    next_row = title_block(ws,
        "Covenant Test Detail — All 8 Credits",
        "Per-covenant actual vs threshold with buffer distance and status classification")

    headers = ["Ticker", "Company", "Sector Tag", "Covenant Name", "Type",
               "Actual", "Threshold", "Operator", "Buffer / Breach", "Status", "Source"]
    col_widths = [8, 26, 14, 32, 10, 12, 12, 10, 14, 14, 45]
    header_row(ws, next_row, headers, col_widths)
    next_row += 1

    new_sectors = {"BHC", "M", "LUMN"}
    for ticker, case in CASE_REGISTRY.items():
        results = evaluate_package(case["covenants"])
        is_new = ticker in new_sectors
        for r, cov in zip(results, case["covenants"]):
            actual_str = f"${r.actual:.0f}M" if r.unit == "$M" else f"{r.actual:.2f}x"
            thresh_str = f"${r.threshold:.0f}M" if r.unit == "$M" else f"{r.threshold:.2f}x"
            buf = r.buffer_pct * 100
            buf_str = f"▼ {abs(buf):.1f}% BREACH" if buf < 0 else f"+{buf:.1f}% buffer"

            row_data = [
                ticker, case["name"][:26], ("★ NEW" if is_new else "original"),
                cov["name"], cov["type"].upper(),
                actual_str, thresh_str,
                "≤ threshold" if cov["operator"] == "lte" else "≥ threshold",
                buf_str, r.status, cov.get("source", "")
            ]
            ws.row_dimensions[next_row].height = 18
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.border = border_thin()
                cell.alignment = left_mid() if col_idx in (2, 4, 11) else center()
                if col_idx == 10:
                    cell.fill = status_fill(r.status)
                    cell.font = status_font(r.status)
                elif col_idx == 9:
                    cell.fill = fill(RED_LIGHT) if buf < 0 else fill(GRN_LIGHT if buf > 10 else AMB_LIGHT)
                    cell.font = font(bold=True,
                        color=(RED if buf < 0 else (GREEN if buf > 10 else AMBER)))
                elif col_idx == 3 and is_new:
                    cell.fill = fill("C8DBFF")
                    cell.font = font(bold=True, color=BLUE)
                else:
                    cell.fill = fill(NEWBLUE if is_new else (WHITE if next_row % 2 == 0 else LGRAY))
                    cell.font = font(bold=(col_idx == 1))
            next_row += 1
        # spacer between companies
        next_row += 0


# ─── SHEET 3: Debt Stack ──────────────────────────────────────────────────────
def build_debt_stack(wb):
    ws = wb.create_sheet("📋 Debt Stack")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    next_row = title_block(ws,
        "Debt Stack — Tranche-Level Detail  ·  All 8 Credits",
        "Seniority · Face Value · Maturity · Cross-Default Clause · Status")

    headers = ["Ticker", "Company", "Tranche", "Face Value ($M)",
               "Seniority", "Rate", "Maturity", "Lender", "Cross-Default?", "Status"]
    col_widths = [8, 26, 42, 16, 28, 24, 12, 28, 15, 12]
    header_row(ws, next_row, headers, col_widths)
    next_row += 1

    new_sectors = {"BHC", "M", "LUMN"}
    for ticker, case in CASE_REGISTRY.items():
        is_new = ticker in new_sectors
        for t in case["debt_stack"]:
            xd = t["cross_default_clause"]
            status = t["status"].upper()
            row_data = [
                ticker, case["name"][:26], t["tranche"],
                t["face_value_usd_m"], t["seniority"],
                t["rate"], t["maturity"], t["lender"],
                "✓ YES" if xd else "— NO", status
            ]
            ws.row_dimensions[next_row].height = 18
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col_idx, value=val)
                cell.border = border_thin()
                cell.alignment = left_mid() if col_idx in (3, 5, 8) else center()
                if col_idx == 10:
                    cell.fill = status_fill(status)
                    cell.font = status_font(status)
                elif col_idx == 9:
                    if xd:
                        cell.fill = fill(RED_LIGHT)
                        cell.font = font(bold=True, color=RED)
                    else:
                        cell.fill = fill(LGRAY)
                        cell.font = font(color=MGRAY)
                elif col_idx == 4:
                    cell.font = font(bold=True, color=DGRAY)
                    cell.number_format = "#,##0"
                else:
                    cell.fill = fill(NEWBLUE if is_new else (WHITE if next_row % 2 == 0 else LGRAY))
                    cell.font = font(bold=(col_idx == 1))
            next_row += 1

    # Total row
    ws.row_dimensions[next_row].height = 22
    total_debt = sum(t["face_value_usd_m"] for case in CASE_REGISTRY.values()
                     for t in case["debt_stack"])
    for c in range(1, 11):
        cell = ws.cell(row=next_row, column=c)
        cell.fill = fill(NAVY)
        cell.border = border_thin()
    ws.cell(row=next_row, column=1, value="TOTAL").font = font(bold=True, color=WHITE)
    ws.cell(row=next_row, column=1).fill = fill(NAVY)
    ws.cell(row=next_row, column=3, value="All Tranches — 8 Credits").font = font(color="B0C4DE")
    ws.cell(row=next_row, column=3).fill = fill(NAVY)
    total_cell = ws.cell(row=next_row, column=4, value=total_debt)
    total_cell.fill = fill(NAVY)
    total_cell.font = font(bold=True, color=WHITE, size=11)
    total_cell.alignment = center()
    total_cell.number_format = "#,##0"


# ─── SHEET 4: Stress Scenarios ─────────────────────────────────────────────────
def build_stress_sheet(wb):
    ws = wb.create_sheet("⚡ Stress Scenarios")
    ws.sheet_view.showGridLines = False
    next_row = title_block(ws,
        "Macro Stress Sensitivity — NLR Under EBITDA Compression",
        "Rows = EBITDA compression shock  ·  Columns = each company  ·  Value = stressed NLR  ·  Red = Breach")

    # Build a single stress scenario: EBITDA shocks at 0%, -10%, -20%, -30%, -40%, -50%
    shocks = [0, -10, -20, -30, -40, -50]
    tickers_ordered = ["BHC", "M", "LUMN", "CHTR", "WBA", "PARA", "HCA", "ATUS"]

    # Header
    ws.row_dimensions[next_row].height = 16
    ws.cell(row=next_row, column=1, value="EBITDA Shock →").fill = fill(NAVY)
    ws.cell(row=next_row, column=1).font = font(bold=True, color=WHITE, size=9)
    ws.cell(row=next_row, column=1).alignment = center()
    ws.cell(row=next_row, column=1).border = border_thin()

    for c_idx, ticker in enumerate(tickers_ordered, 2):
        case = CASE_REGISTRY[ticker]
        is_new = ticker in {"BHC", "M", "LUMN"}
        label = f"{ticker}{'  ★' if is_new else ''}\n(thresh {case['covenants'][0]['threshold']:.1f}x)"
        cell = ws.cell(row=next_row, column=c_idx, value=label)
        cell.fill = fill(MIDBLUE if is_new else NAVY)
        cell.font = font(bold=True, color=WHITE, size=9)
        cell.alignment = center()
        cell.border = border_thin()
        ws.column_dimensions[get_column_letter(c_idx)].width = 14
    ws.column_dimensions["A"].width = 18
    next_row += 1

    for shock in shocks:
        ws.row_dimensions[next_row].height = 20
        shock_label = f"Base" if shock == 0 else f"EBITDA {shock}%"
        label_cell = ws.cell(row=next_row, column=1, value=shock_label)
        label_cell.fill = fill(LGRAY)
        label_cell.font = font(bold=True, color=DGRAY)
        label_cell.alignment = center()
        label_cell.border = border_thin()

        for c_idx, ticker in enumerate(tickers_ordered, 2):
            case = CASE_REGISTRY[ticker]
            thresh = case["covenants"][0]["threshold"]
            ebitda_stressed = case["ebitda_ltm"] * (1 + shock / 100)
            nd = case["total_debt"] - case["cash"]
            nlr_stressed = nd / ebitda_stressed if ebitda_stressed > 0 else 999

            breached = nlr_stressed > thresh
            near = (thresh - nlr_stressed) / thresh < 0.10 and not breached

            val_cell = ws.cell(row=next_row, column=c_idx, value=round(nlr_stressed, 2))
            val_cell.number_format = '0.00"x"'
            val_cell.alignment = center()
            val_cell.border = border_thin()

            if breached:
                val_cell.fill = fill(RED_LIGHT)
                val_cell.font = font(bold=True, color=RED)
            elif near:
                val_cell.fill = fill(AMB_LIGHT)
                val_cell.font = font(bold=True, color=AMBER)
            else:
                val_cell.fill = fill(GRN_LIGHT)
                val_cell.font = font(color=GREEN)

        next_row += 1

    # Key insight rows
    next_row += 1
    insights = [
        ("BHC", "Already in breach. $3.5B TLB matures Jun 2025 — refinancing is acute.", RED),
        ("M",   "Comfortable headroom — but Arkhouse LBO bid would push pro-forma NLR to ~5x.", AMBER),
        ("LUMN","Marginally breached NLR + deep FCCR breach. 2nd restructuring base-case.", RED),
        ("ATUS","Deep breach. No stress scenario improves position. Distressed outcome likely.", RED),
        ("HCA", "Healthiest credit. Only breach under 40% EBITDA compression + 20% debt increase.", GREEN),
    ]
    for ticker, insight, color in insights:
        ws.row_dimensions[next_row].height = 18
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=9)
        cell = ws.cell(row=next_row, column=1,
                       value=f"  {ticker}:  {insight}")
        cell.fill = fill(RED_LIGHT if color == RED else (AMB_LIGHT if color == AMBER else GRN_LIGHT))
        cell.font = font(bold=False, color=color, size=9)
        cell.alignment = left_mid()
        cell.border = border_thin()
        next_row += 1


# ─── SHEET 5: About ────────────────────────────────────────────────────────────
def build_about(wb):
    ws = wb.create_sheet("ℹ️ About")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 72

    next_row = title_block(ws,
        "DDCSE v3.0 — Dynamic Debt Covenant Surveillance Engine",
        "About This Workbook  ·  Data Sources  ·  Methodology  ·  Disclaimer")

    rows = [
        ("Engine", "Dynamic Debt Covenant Surveillance Engine (DDCSE) v3.0"),
        ("Author", "Anklesh Rawat  ·  github.com/DogInfantry"),
        ("Repository", "github.com/DogInfantry/dynamic-debt-covenant-surveillance-frame"),
        ("Generated", date.today().strftime("%B %d, %Y")),
        ("Credits Covered", "8 companies across 6 sectors (Media/Cable, Healthcare Retail, Entertainment, Hospitals, Legacy Telecom, Specialty Pharma, Dept Store Retail, Enterprise Fiber Telecom)"),
        ("New in v3.0", "BHC (Bausch Health — Specialty Pharma)  ·  M (Macy's — Dept Store Retail)  ·  LUMN (Lumen Technologies — Enterprise Fiber Telecom)"),
        ("", ""),
        ("SHEET GUIDE", ""),
        ("📊 Portfolio Dashboard", "One-row-per-company overview with RAG covenant status, NLR vs threshold, buffer %, severity score"),
        ("🔍 Covenant Detail", "All four covenant tests (NLR, ICR, FCCR, Min. Liquidity) for every company with actual vs threshold"),
        ("📋 Debt Stack", "Tranche-level debt breakdown — seniority, face value, maturity, lender, cross-default clause"),
        ("⚡ Stress Scenarios", "EBITDA compression stress table showing stressed NLR at -0% to -50% EBITDA for each company"),
        ("", ""),
        ("DATA SOURCES", ""),
        ("CHTR", "10-K FY2023 (Feb 9, 2024) · SEC CIK 0001091882 · Charter Operating Credit Agreement (2021) §7.1"),
        ("WBA",  "10-K FY2023 (Oct 12, 2023) · SEC CIK 0001618921 · WBA Credit Agreement (2021) §6.1"),
        ("PARA", "10-K FY2023 (Feb 27, 2024) · SEC CIK 0000813828 · PARA Credit Agreement (2021) §7.01"),
        ("HCA",  "10-K FY2023 (Feb 23, 2024) · SEC CIK 0000860731 · HCA Senior Secured Credit Facilities (2023) §7.1"),
        ("ATUS", "10-K FY2023 (Mar 5, 2024) · SEC CIK 0001672013 · CSC Holdings Credit Agreement (2019) §7.1"),
        ("BHC ★NEW", "10-K FY2023 (Mar 6, 2024) · SEC CIK 0000885590 · BHC Secured Credit Agreement (2022) §7.1"),
        ("M ★NEW",   "10-K FY2023 (Feb 3, 2024) · SEC CIK 0000794367 · Macy's Credit Agreement (2023 amendment) §7.1"),
        ("LUMN ★NEW","10-K FY2023 (Mar 5, 2024) · SEC CIK 0000018926 · Lumen Credit Agreement (2022 amendment) §7.1"),
        ("", ""),
        ("DISCLAIMER", "This workbook is for analytical and portfolio monitoring purposes only. "
         "All financial data is sourced from publicly available SEC EDGAR filings as noted above. "
         "This does not constitute investment advice, a solicitation to buy or sell securities, "
         "or a credit opinion. Figures are as of the filing dates noted and may not reflect "
         "subsequent developments, amendments, or waivers. Not investment advice."),
    ]

    for label, value in rows:
        ws.row_dimensions[next_row].height = 20 if label not in ("", "SHEET GUIDE", "DATA SOURCES", "DISCLAIMER") else 22
        lc = ws.cell(row=next_row, column=1, value=label)
        vc = ws.cell(row=next_row, column=2, value=value)
        lc.border = border_thin()
        vc.border = border_thin()
        lc.alignment = left_mid()
        vc.alignment = left_mid()

        if label in ("SHEET GUIDE", "DATA SOURCES", "DISCLAIMER", "Engine"):
            lc.fill = fill(BLUE); lc.font = font(bold=True, color=WHITE, size=9)
            vc.fill = fill(BLUE); vc.font = font(bold=True, color=WHITE, size=9)
        elif label == "":
            lc.fill = fill(LGRAY); vc.fill = fill(LGRAY)
        elif "★NEW" in label:
            lc.fill = fill(NEWBLUE); lc.font = font(bold=True, color=BLUE)
            vc.fill = fill(NEWBLUE)
        elif label in ("New in v3.0",):
            lc.fill = fill(NEWBLUE); lc.font = font(bold=True, color=BLUE)
            vc.fill = fill("D8E8FF"); vc.font = font(bold=False, color=BLUE)
        else:
            lc.fill = fill(LGRAY if next_row % 2 == 0 else WHITE)
            vc.fill = fill(LGRAY if next_row % 2 == 0 else WHITE)
            lc.font = font(bold=True, color=DGRAY, size=9)
            vc.font = font(color=DGRAY, size=9)
        next_row += 1

    ws.print_area = f"A1:B{next_row}"


# ─── MAIN ──────────────────────────────────────────────────────────────────────
def build_excel(output_path):
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_dashboard(wb)
    build_covenant_detail(wb)
    build_debt_stack(wb)
    build_stress_sheet(wb)
    build_about(wb)

    wb.save(output_path)
    print(f"✓ Excel written: {output_path}")


if __name__ == "__main__":
    out = "reports_for_stakeholders/DDCSE_Covenant_Surveillance_v3.xlsx"
    os.makedirs("reports_for_stakeholders", exist_ok=True)
    build_excel(out)
